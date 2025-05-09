# utils/excel_writer.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

def style_sheet(ws):
    """Applies general styling to a worksheet."""
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    content_alignment = Alignment(vertical="middle", wrap_text=True, shrink_to_fit=False) # Allow wrap_text to expand row height
    
    thin_border_side = Side(border_style="thin", color="000000")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    # Style content cells and set column widths
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.alignment = content_alignment
            cell.border = cell_border
            # Add padding (OpenPyxl doesn't directly support padding, achieve via alignment and cell size)
    
    # Auto-fit column widths with some padding
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 5) * 1.2 # Add padding and factor
        if adjusted_width > 70: # Max width
             adjusted_width = 70
        if column_letter == 'A' and ws.title in ["Email", "LinkedIn", "FaceBook"]: # Version #
            ws.column_dimensions[column_letter].width = 10
            for cell in ws[column_letter]:
                cell.alignment = Alignment(horizontal="center", vertical="middle")
        else:
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Ensure row height adjusts to wrapped text
        for row_dim in ws.row_dimensions.values():
            row_dim.auto_size = True # This might not work as expected, manual height or wrap_text is key

    # Ensure rows can expand
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = None # Let it auto-adjust based on content and wrap_text

def create_excel_report(all_ad_data, company_name, lead_objective_user_selection):
    """Creates an Excel report from the generated ad data."""
    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet

    # Email Page
    if 'email' in all_ad_data and all_ad_data['email']:
        ws_email = wb.create_sheet("Email")
        email_headers = ["Version #", "Objective", "Headline", "Subject Line", "Body", "CTA"]
        ws_email.append(email_headers)
        for i, ad in enumerate(all_ad_data['email']):
            ws_email.append([
                i + 1,
                "Demand Capture", # Fixed as per spec
                ad.get("headline", ""),
                ad.get("subject_line", ""),
                ad.get("body", ""),
                ad.get("cta", "")
            ])
        style_sheet(ws_email)

    # LinkedIn Page
    if 'linkedin' in all_ad_data and all_ad_data['linkedin']:
        ws_linkedin = wb.create_sheet("LinkedIn")
        linkedin_headers = ["Version #", "Ad Name", "Objective", "Introductory Text", "Image Copy", "Headline", "Destination", "CTA Button"]
        ws_linkedin.append(linkedin_headers)
        version_counters = {"Brand Awareness": 0, "Demand Gen": 0, "Demand Capture": 0}
        for ad in all_ad_data['linkedin']:
            obj = ad.get("objective_type", "Unknown")
            version_counters[obj] += 1
            ws_linkedin.append([
                version_counters[obj],
                ad.get("ad_name", ""),
                obj,
                ad.get("introductory_text", ""),
                ad.get("image_copy", ""),
                ad.get("headline", ""),
                ad.get("destination_url", ""),
                ad.get("cta_button", "")
            ])
        style_sheet(ws_linkedin)

    # FaceBook Page
    if 'facebook' in all_ad_data and all_ad_data['facebook']:
        ws_facebook = wb.create_sheet("FaceBook")
        facebook_headers = ["Version #", "Ad Name", "Objective", "Primary Text", "Image Copy", "Headline", "Link Description", "Destination", "CTA Button"]
        ws_facebook.append(facebook_headers)
        version_counters = {"Brand Awareness": 0, "Demand Gen": 0, "Demand Capture": 0}
        for ad in all_ad_data['facebook']:
            obj = ad.get("objective_type", "Unknown")
            version_counters[obj] += 1
            ws_facebook.append([
                version_counters[obj],
                ad.get("ad_name", ""),
                obj,
                ad.get("primary_text", ""),
                ad.get("image_copy", ""),
                ad.get("headline", ""),
                ad.get("link_description", ""),
                ad.get("destination_url", ""),
                ad.get("cta_button", "")
            ])
        style_sheet(ws_facebook)

    # Google Search Page
    if 'google_search' in all_ad_data and all_ad_data['google_search']:
        ws_gsearch = wb.create_sheet("Google Search")
        gsearch_headers = ["Headline", "Description"]
        ws_gsearch.append(gsearch_headers)
        
        headlines = all_ad_data['google_search'].get("headlines", [])
        descriptions = all_ad_data['google_search'].get("descriptions", [])
        
        max_rows = max(len(headlines), len(descriptions))
        for i in range(max_rows):
            h = headlines[i] if i < len(headlines) else ""
            d = descriptions[i] if i < len(descriptions) else ""
            # Only write description if it's one of the first 4, corresponding to headlines
            if i < len(headlines) and i < 4: # Descriptions are fewer
                 ws_gsearch.append([h,d])
            elif i < len(headlines): # Only headline
                 ws_gsearch.append([h,""])
            # This logic needs to match the structure: 15H, 4D.
            # The prompt asks for 15H and 4D. So we should list all 15H, and the 4D alongside the first 4H.
        
        # Revised Google Search population
        ws_gsearch.delete_rows(2, ws_gsearch.max_row) # Clear previous attempts
        for i in range(15): # 15 Headlines
            headline_text = headlines[i] if i < len(headlines) else ""
            description_text = descriptions[i] if i < len(descriptions) and i < 4 else "" # Only 4 descriptions
            ws_gsearch.append([headline_text, description_text])
        style_sheet(ws_gsearch)


    # Google Display Page
    if 'google_display' in all_ad_data and all_ad_data['google_display']:
        ws_gdisplay = wb.create_sheet("Google Display")
        gdisplay_headers = ["Headline", "Description"]
        ws_gdisplay.append(gdisplay_headers)
        
        headlines = all_ad_data['google_display'].get("headlines", [])
        descriptions = all_ad_data['google_display'].get("descriptions", [])

        # Revised Google Display population (5 headlines, 5 descriptions)
        for i in range(5):
            headline_text = headlines[i] if i < len(headlines) else ""
            description_text = descriptions[i] if i < len(descriptions) else ""
            ws_gdisplay.append([headline_text, description_text])
        style_sheet(ws_gdisplay)

    # Reasoning Page
    if 'reasoning' in all_ad_data:
        ws_reasoning = wb.create_sheet("Reasoning")
        reasoning_content = all_ad_data['reasoning']
        
        ws_reasoning.append(["Section", "Content"])
        
        if reasoning_content.get('url_summary'):
            ws_reasoning.append(["URL Context Summary", reasoning_content['url_summary']])
        if reasoning_content.get('additional_summary'):
            ws_reasoning.append(["Additional Context Summary", reasoning_content['additional_summary']])
        if reasoning_content.get('downloadable_summary'):
            ws_reasoning.append(["Downloadable Material Summary", reasoning_content['downloadable_summary']])
        if reasoning_content.get('ai_reasoning'):
            ws_reasoning.append(["AI's Reasoning & Thought Process", reasoning_content['ai_reasoning']])
        
        style_sheet(ws_reasoning)
        # Specific styling for reasoning page if needed (e.g., merge cells for content)
        for row in ws_reasoning.iter_rows(min_row=2, max_col=2):
            row[0].font = Font(bold=True) # Make section titles bold
            row[1].alignment = Alignment(vertical="top", wrap_text=True) # Align content top for long texts
        ws_reasoning.column_dimensions['B'].width = 100 # Wider column for content

    # Save to a BytesIO object
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    # Company name from URL, simple extraction
    filename_company_part = company_name.replace("www.", "").split('.')[0] if company_name else "company"
    filename = f"{filename_company_part}_{lead_objective_user_selection.lower().replace(' ', '_')}.xlsx"
    
    return excel_bytes, filename